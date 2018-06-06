from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.models import User, Group

from izpiti.models import *
from student.models import *
from .forms import *
from .models import *
import datetime

from openpyxl import *

def vpisi_studenta(student, leto, smer ,letnik, a_vs2):
    
    katarina_vpis = Vpis(student=student, 
    studijsko_leto= leto, 
    studijski_program=smer, 
    letnik=letnik, 
    vrsta_vpisa=a_vv1,
    nacin_studija=a_ns1,
    oblika_studija = a_oblika_on_site,
    potrjen = True, 
    dokoncan_vpis = True,
    vrsta_studija=a_vs2)
    
    katarina_vpis.save()

    p_predmetiStudentpPrimoz0_old = PredmetiStudenta()
    p_predmetiStudentpPrimoz0_old.save()
    p_predmetiStudentpPrimoz0_old.vpis = katarina_vpis

    if smer == a_UNI:
        if letnik == a_1Letnik:
            p_predmetiStudentpPrimoz0_old.predmeti.add(*UNI_PREDMETI_PRVI)
        elif letnik == a_2Letnik:
            p_predmetiStudentpPrimoz0_old.predmeti.add(*UNI_PREDMETI_DRUGI)
        elif letnik == a_3Letnik:
            if student == student14 or student == student15:
                p_predmetiStudentpPrimoz0_old.predmeti.add(*UNI_PREDMETI_TRETJI_2)
            elif student == student16 or student == student17 or student == student18:
                p_predmetiStudentpPrimoz0_old.predmeti.add(*UNI_PREDMETI_TRETJI_3)
            elif student == student19 or student == student20:
                p_predmetiStudentpPrimoz0_old.predmeti.add(*UNI_PREDMETI_TRETJI_4)    
            
            else:
                p_predmetiStudentpPrimoz0_old.predmeti.add(*UNI_PREDMETI_TRETJI_1)            
    else:
        p_predmetiStudentpPrimoz0_old.predmeti.add(*VS_PREDMETI_DRUGI)

    



    
    p_predmetiStudentpPrimoz0_old.save()

def naredi_studenta(ime,priimek,vpisna,email):

    katarina = Student(vpisna_stevilka = vpisna, 
    emso = "1511996500207", 
    priimek=priimek, 
    ime=ime, 
    naslov_stalno_bivalisce="Dunajska cesta 110", 
    drzava=a_slo, 
    drzava_rojstva = a_slo, 
    posta= a_sklPosta, 
    obcina=a_sklObcina,
    obcina_rojstva= a_ljObcina, 
    telefon="031866686", 
    email=email)
    katarina.save()
    user, created = User.objects.get_or_create(username=email[:6], email=email)
    user.first_name = ime
    user.last_name = priimek
        
    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        prof_group, status = Group.objects.get_or_create(name='students') 
        prof_group.user_set.add(user)

    user.save()
    return katarina



def narediIzvedboSTremi(predmetnik, prof1,prof2, prof3):
    
    prof1.predmeti.add(predmetnik.predmet)
    prof2.predmeti.add(predmetnik.predmet)
    prof3.predmeti.add(predmetnik.predmet)
    izvedba = IzvedbaPredmeta(predmet = predmetnik.predmet,
                         studijsko_leto= predmetnik.studijsko_leto, 
                         ucitelj_1= prof1,
                         ucitelj_2 = prof2,
                         ucitelj_3 = prof3)
    izvedba.save()
    prof1.save()
    prof2.save()
    prof3.save()
    return

def narediIzvedboZDvemi(predmetnik, prof1,prof2):
    
    prof1.predmeti.add(predmetnik.predmet)
    prof2.predmeti.add(predmetnik.predmet)
    izvedba = IzvedbaPredmeta(predmet = predmetnik.predmet,
                         studijsko_leto= predmetnik.studijsko_leto, 
                         ucitelj_1= prof1,
                         ucitelj_2 = prof2)
    izvedba.save()
    prof1.save()
    prof2.save()
    return
#iz ucitelja in predmetnika naredi novo izvedbo predmeta
def narediIzvedbo(predmetnik, prof):

    prof.predmeti.add(predmetnik.predmet)
    izvedba = IzvedbaPredmeta(predmet = predmetnik.predmet, studijsko_leto= predmetnik.studijsko_leto, ucitelj_1= prof)
    izvedba.save()
    prof.save()
    return

def brezSumnikov(beseda):
    a = beseda

    i = 0
    for c in beseda:
        if c == "č":
            a = a[:i] + "c" + a[i+1:]
        elif c == "š":
            a = a[:i] + "s" + a[i+1:]
        elif c == "ž":
            a = a[:i] + "z" + a[i+1:]
        elif c == "ć":
            a = a[:i] + "c" + a[i+1:]
        i = i+1
    return a

def narediProfesorja(imes,priimeks):
    ime = brezSumnikov(imes.lower())
    priimek = brezSumnikov(priimeks.lower())
    if Ucitelj.objects.filter(ime=imes,priimek=priimeks).exists():
        return Ucitelj.objects.filter(ime=imes,priimek=priimeks)[0]

    user, created = User.objects.get_or_create(username=ime +priimek, email=ime+"."+priimek+  "@fri.uni-lj.si")
    user.first_name = imes
    user.last_name = priimeks
        
    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        prof_group, status = Group.objects.get_or_create(name='professors') 
        prof_group.user_set.add(user)

    user.save()

    profesor = Ucitelj(ime = imes, priimek = priimeks, email = ime+"."+priimek+  "@fri.uni-lj.si")
    #print(profesor.email)
    profesor.save()
    return profesor


def uvozi_drzave():
    print('\x1b[6;30;42m' + 'Začetek uvažanja držav ' + '\x1b[0m')
    wb = load_workbook('sifranti/ŠifrantDržav.xlsx')
    sheet = wb.active
    cellsa = sheet['A3':'F251']
    for dvomestna, tromestna, numericna_oz, iso, slovenski, opombaa in cellsa:
        if opombaa.value == None:
            opombaa.value = ""

        a = Drzava(id=numericna_oz.value, 
        dvomestna_koda=dvomestna.value, 
        tromestna_oznaka=tromestna.value,
         iso_naziv=iso.value, 
         slovenski_naziv=slovenski.value,
         opomba=opombaa.value, 
         veljaven=True)

        a.save()
    print('\x1b[6;30;42m' + 'Konec uvažanja držav ' + '\x1b[0m')
        
    return

def uvozi_obcine():
    print('\x1b[6;30;42m' + 'Začetek uvažanja občin ' + '\x1b[0m')
    wb = load_workbook('sifranti/ŠifrantObčin.xlsx')
    sheet = wb.active
    cellsa = sheet['A2':'B213']
    for sifra, imee in cellsa:
        
        a = Obcina(id = sifra.value, ime = imee.value)

        a.save()
    print('\x1b[6;30;42m' + 'Konec uvažanja občin ' + '\x1b[0m')
        
    return

def uvozi_poste():
    print('\x1b[6;30;42m' + 'Začetek uvažanja pošt ' + '\x1b[0m')
    wb = load_workbook('sifranti/ŠifrantPošt.xlsx')
    sheet = wb.active
    cellsa = sheet['B5':'C482']
    for sifra, imee in cellsa:
        
        a = Posta(id = sifra.value, kraj = imee.value)

        a.save()
    print('\x1b[6;30;42m' + 'Konec uvažanja pošt ' + '\x1b[0m')
        
    return

def naredi_bazo(request):
    
    print('\x1b[6;30;42m' + 'Začetek delanja baze ' + '\x1b[0m')

    uvozi_drzave()
    uvozi_obcine()
    uvozi_poste()

    global a_slo
    global a_sklObcina
    global a_ljObcina
    global a_ljPosta
    global a_sklPosta

    a_slo = Drzava.objects.get(id=705)
    a_sklObcina = Obcina.objects.get(id = 122)
    a_ljObcina = Obcina.objects.get(id = 61)
    a_ljPosta = Posta.objects.get(id=1000)
    a_sklPosta = Posta.objects.get(id=4220)
    a_sklPosta.save()

    a_studijskiProgram = StudijskiProgram(id=1000475,sifra="L2",stopnja="C - (predbolonjski) univerzitetni", semestri=9, naziv= "RAČUNAL. IN INFORMATIKA UN")
    a_studijskiProgram.save()
    a_vrstaStudija = VrstaStudija(id=12001,opis="Osnovnošolska izobrazba", nacin_zakljucka="zakljucena osnovna šola", raven_klasius=1)
    a_vrstaStudija.save()
    a_vrstaVpisa = VrstaVpisa(id=1, opis="Prvi vpis v letnik/dodatno leto", mozni_letniki="Vsi letniki in dodatno leto")
    a_vrstaVpisa.save()
    a_nacinStudija = NacinStudija(id=1, opis="redni",ang_opis="full-time")
    a_nacinStudija.save()
    a = OblikaStudija(id=1, opis="na lokaciji", ang_opis="on-site" )
    a.save()
    global a_oblika_on_site
    a_oblika_on_site = a

    #naredi studenta
    primozt = Student(vpisna_stevilka = 63150000,
     emso=1511996500207, 
     ime="Primož", 
     drzava_rojstva=a_slo,
     obcina_rojstva= a_ljObcina,  
     priimek="Trubar",
     naslov_stalno_bivalisce="Kranjska ulica 12",
      drzava= Drzava.objects.filter(pk=4)[0], 
      posta= Posta.objects.filter(pk=1293)[0],obcina= Obcina.objects.filter(pk=1)[0],telefon="040123456",email="pt0000@fri.uni-lj.si")
    primozt.save()

    

    user, created = User.objects.get_or_create(username="student", email="pt0000@fri.uni-lj.si")
    user.first_name = "Primož"
    user.last_name = "Trubar"

    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        ref_group, status = Group.objects.get_or_create(name='students') 
        ref_group.user_set.add(user)
    
    user.save()

    # student 1 letnik
    ivan = Student(vpisna_stevilka = 63180000, emso=1511996500207, ime="Ivan", drzava_rojstva=a_slo,obcina_rojstva= a_ljObcina,  priimek="Cankar",naslov_stalno_bivalisce="Na klancu 29", drzava= Drzava.objects.filter(pk=4)[0], posta= Posta.objects.filter(pk=1293)[0],obcina= Obcina.objects.filter(pk=1)[0],telefon="040123456",email="ic1111@fri.uni-lj.si")
    ivan.save()

    user, created = User.objects.get_or_create(username="ivan", email="ic1111@fri.uni-lj.si")
    user.first_name = "Ivan"
    user.last_name = "Cankar"

    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        ref_group, status = Group.objects.get_or_create(name='students') 
        ref_group.user_set.add(user)
    
    user.save()

    # student 2 letnik
    drago = Student(vpisna_stevilka = 63170000, emso=1511996500207, ime="Drago", drzava_rojstva=a_slo,obcina_rojstva= a_ljObcina,  priimek="Furja",naslov_stalno_bivalisce="Ropotoče 28", drzava= Drzava.objects.filter(pk=4)[0], posta= Posta.objects.filter(pk=1293)[0],obcina= Obcina.objects.filter(pk=1)[0],telefon="040123456",email="df2222@fri.uni-lj.si")
    drago.save()

    user, created = User.objects.get_or_create(username="drago", email="df2222@fri.uni-lj.si")
    user.first_name = "Drago"
    user.last_name = "Furja"

    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        ref_group, status = Group.objects.get_or_create(name='students') 
        ref_group.user_set.add(user)
    
    user.save()


    # student 3 letnik brez izbire
    martin = Student(vpisna_stevilka = 63160000, emso=1511996500207, ime="Martin", drzava_rojstva=a_slo,obcina_rojstva= a_ljObcina,  priimek="Luter",naslov_stalno_bivalisce="Dunajska 124", drzava= Drzava.objects.filter(pk=4)[0], posta= Posta.objects.filter(pk=1293)[0],obcina= Obcina.objects.filter(pk=1)[0],telefon="040123456",email="ml3333@fri.uni-lj.si")
    martin.save()

    user, created = User.objects.get_or_create(username="martin", email="ml3333@fri.uni-lj.si")
    user.first_name = "Martin"
    user.last_name = "Luter"

    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        ref_group, status = Group.objects.get_or_create(name='students') 
        ref_group.user_set.add(user)
    
    user.save()

    # student 3 letnik brez izbire
    tine = Student(vpisna_stevilka = 63160001, emso=1511996500207, ime="Tine", drzava_rojstva=a_slo,obcina_rojstva= a_ljObcina,  priimek="Prejovc",naslov_stalno_bivalisce="Tavčarjeva 4", drzava= Drzava.objects.filter(pk=4)[0], posta= Posta.objects.filter(pk=1293)[0],obcina= Obcina.objects.filter(pk=1)[0],telefon="040123456",email="tp4444@fri.uni-lj.si")
    tine.save()

    user, created = User.objects.get_or_create(username="tine", email="tp4444@fri.uni-lj.si")
    user.first_name = "Tine"
    user.last_name = "Prejovc"

    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        ref_group, status = Group.objects.get_or_create(name='students') 
        ref_group.user_set.add(user)
    
    user.save()




    #naredi referenta
    user, created = User.objects.get_or_create(username="referentka", email="referentka@fri.uni-lj.si")
    user.first_name = "Tatjana"
    user.last_name = "Novak"
        
    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        ref_group, status = Group.objects.get_or_create(name='referent') 
        ref_group.user_set.add(user)

    global a_1Letnik
    a_1Letnik = Letnik(ime="1.")
    a_1Letnik.save()
    global a_2Letnik
    a_2Letnik = Letnik(ime="2.")
    a_2Letnik.save()
    global a_3Letnik
    a_3Letnik= Letnik(ime="3.")
    a_3Letnik.save()


    a_stud = StudijskiProgram(id=1000475,sifra="L2",stopnja="C - (predbolonjski) univerzitetni", semestri=9, naziv= "RAČUNAL. IN INFORMATIKA UN")
    a_stud.save()
    a_stud2 = StudijskiProgram(id=1000471,sifra="L1",stopnja="L - druga stopnja: magistrski", semestri=4, naziv= "RAČUNALN. IN INFORM. MAG II.ST")
    a_stud2.save()

    global a_UNI
    global a_VS

    a_UNI = StudijskiProgram(id=1000468 ,sifra="VT",stopnja="K - prva stopnja: univerzitetni", semestri=6, naziv= "RAČUNALN. IN INFORM. UN-I.ST")
    a_UNI.save()
    a_VS = StudijskiProgram(id=1000470 ,sifra="VU",stopnja="J - prva stopnja: visokošolski strokovni", semestri=6, naziv= "RAČUNALN. IN INFORM. VS-I.ST")
    a_VS.save()

    a_15_16 = StudijskoLeto(ime = "2015/2016")
    a_15_16.save()
    a_16_17 = StudijskoLeto(ime = "2016/2017")
    a_16_17.save()
    a_17_18 = StudijskoLeto(ime = "2017/2018")
    a_17_18.save()
    a_18_19 = StudijskoLeto(ime = "2018/2019")
    a_18_19.save()

    a_vilijan = Ucitelj(ime = "Viljan", priimek = "Mahnič", email = "viljan.mahnic@fri.uni-lj.si")
    a_vilijan.save()

    print('\x1b[6;30;42m' + 'Uvažanje predmetov leto 2018/2019 ' + '\x1b[0m')
    vsi_predmeti()
    print('\x1b[6;30;42m' + 'Uvažanje predmetov leto 2017/2018 ' + '\x1b[0m')
    vsi_predmeti2(a_17_18)
    print('\x1b[6;30;42m' + 'Uvažanje predmetov leto 2016/2017 ' + '\x1b[0m')
    vsi_predmeti2(a_16_17)
    print('\x1b[6;30;42m' + 'Uvažanje predmetov leto 2015/2016 ' + '\x1b[0m')
    vsi_predmeti2(a_15_16)

    print('\x1b[6;30;42m' + 'Ostali ' + '\x1b[0m')

    global UNI_PREDMETI_PRVI
    UNI_PREDMETI_PRVI = []
    UNI_PREDMETI_PRVI.append(Predmet.objects.get(id=63277))
    UNI_PREDMETI_PRVI.append(Predmet.objects.get(id=63278))
    UNI_PREDMETI_PRVI.append(Predmet.objects.get(id=63203))
    UNI_PREDMETI_PRVI.append(Predmet.objects.get(id=63205))
    UNI_PREDMETI_PRVI.append(Predmet.objects.get(id=63204))
    UNI_PREDMETI_PRVI.append(Predmet.objects.get(id=63202))
    UNI_PREDMETI_PRVI.append(Predmet.objects.get(id=63207))
    UNI_PREDMETI_PRVI.append(Predmet.objects.get(id=63215))
    UNI_PREDMETI_PRVI.append(Predmet.objects.get(id=63209))
    UNI_PREDMETI_PRVI.append(Predmet.objects.get(id=63212))

    global UNI_PREDMETI_DRUGI
    UNI_PREDMETI_DRUGI = []
    UNI_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63213))
    UNI_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63279))
    UNI_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63208))
    UNI_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63218))
    UNI_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63283))
    UNI_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63216))
    UNI_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63280))
    UNI_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63217))
    UNI_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63220))
    UNI_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63219))

    # drugi are actually prvi
    global VS_PREDMETI_DRUGI
    VS_PREDMETI_DRUGI = []
    VS_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63710))
    VS_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63709))
    VS_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63708))
    VS_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63702))
    VS_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63707))
    VS_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63706))
    VS_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63703))
    VS_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63705))
    VS_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63704))
    VS_PREDMETI_DRUGI.append(Predmet.objects.get(id = 63701))

    global UNI_PREDMETI_TRETJI_1
    UNI_PREDMETI_TRETJI_1 = []
    UNI_PREDMETI_TRETJI_1.append(Predmet.objects.get(id = 63214))
    UNI_PREDMETI_TRETJI_1.append(Predmet.objects.get(id = 63248))
    UNI_PREDMETI_TRETJI_1.append(Predmet.objects.get(id = 63281))
    UNI_PREDMETI_TRETJI_1.append(Predmet.objects.get(id = 63265))
    UNI_PREDMETI_TRETJI_1.append(Predmet.objects.get(id = 63264))
    UNI_PREDMETI_TRETJI_1.append(Predmet.objects.get(id = 63263))
    UNI_PREDMETI_TRETJI_1.append(Predmet.objects.get(id = 63271))
    UNI_PREDMETI_TRETJI_1.append(Predmet.objects.get(id = 63270))
    UNI_PREDMETI_TRETJI_1.append(Predmet.objects.get(id = 63269))
    UNI_PREDMETI_TRETJI_1.append(Predmet.objects.get(id = 63284))
    UNI_PREDMETI_TRETJI_1.append(Predmet.objects.get(id = 63222))

    global UNI_PREDMETI_TRETJI_2
    UNI_PREDMETI_TRETJI_2 = []
    UNI_PREDMETI_TRETJI_2.append(Predmet.objects.get(id = 63214))
    UNI_PREDMETI_TRETJI_2.append(Predmet.objects.get(id = 63248))
    UNI_PREDMETI_TRETJI_2.append(Predmet.objects.get(id = 63281))
    UNI_PREDMETI_TRETJI_2.append(Predmet.objects.get(id = 63249))
    UNI_PREDMETI_TRETJI_2.append(Predmet.objects.get(id = 63251))
    UNI_PREDMETI_TRETJI_2.append(Predmet.objects.get(id = 63252))
    UNI_PREDMETI_TRETJI_2.append(Predmet.objects.get(id = 63253))
    UNI_PREDMETI_TRETJI_2.append(Predmet.objects.get(id = 63266))
    UNI_PREDMETI_TRETJI_2.append(Predmet.objects.get(id = 63271))
    UNI_PREDMETI_TRETJI_2.append(Predmet.objects.get(id = 63257))

    global UNI_PREDMETI_TRETJI_3
    UNI_PREDMETI_TRETJI_3 = []
    UNI_PREDMETI_TRETJI_3.append(Predmet.objects.get(id = 63214))
    UNI_PREDMETI_TRETJI_3.append(Predmet.objects.get(id = 63248))
    UNI_PREDMETI_TRETJI_3.append(Predmet.objects.get(id = 63281))
    UNI_PREDMETI_TRETJI_3.append(Predmet.objects.get(id = 63265))
    UNI_PREDMETI_TRETJI_3.append(Predmet.objects.get(id = 63264))
    UNI_PREDMETI_TRETJI_3.append(Predmet.objects.get(id = 63263))
    UNI_PREDMETI_TRETJI_3.append(Predmet.objects.get(id = 63256))
    UNI_PREDMETI_TRETJI_3.append(Predmet.objects.get(id = 63255))
    UNI_PREDMETI_TRETJI_3.append(Predmet.objects.get(id = 63254))
    UNI_PREDMETI_TRETJI_3.append(Predmet.objects.get(id = 63214))
    UNI_PREDMETI_TRETJI_3.append(Predmet.objects.get(id = 63746))

    global UNI_PREDMETI_TRETJI_4
    UNI_PREDMETI_TRETJI_4 = []
    UNI_PREDMETI_TRETJI_4.append(Predmet.objects.get(id = 63214))
    UNI_PREDMETI_TRETJI_4.append(Predmet.objects.get(id = 63248))
    UNI_PREDMETI_TRETJI_4.append(Predmet.objects.get(id = 63281))
    UNI_PREDMETI_TRETJI_4.append(Predmet.objects.get(id = 63250))
    UNI_PREDMETI_TRETJI_4.append(Predmet.objects.get(id = 63249))
    UNI_PREDMETI_TRETJI_4.append(Predmet.objects.get(id = 63253))
    UNI_PREDMETI_TRETJI_4.append(Predmet.objects.get(id = 63259))
    UNI_PREDMETI_TRETJI_4.append(Predmet.objects.get(id = 63258))
    UNI_PREDMETI_TRETJI_4.append(Predmet.objects.get(id = 63257))
    UNI_PREDMETI_TRETJI_4.append(Predmet.objects.get(id = 63267))
    
    




    #
    a_teh = Predmet.objects.get(ime = "Tehnologija programske opreme")

    a_obl = Predmet.objects.get(ime = "Osnove oblikovanja")
    
    a_ep = Predmet.objects.get(ime = "Ekonomika in podjetništvo")

    a_oim = Predmet.objects.get(ime = "Organizacija in management")

    a_P1 = Predmet.objects.filter(ime = "Programiranje 1")[0]

    a = Predmet.objects.filter(ime = "Programiranje 2")[0]

    a_aps1 = Predmet.objects.get(ime = "Algoritmi in podatkovne strukture 1")

    
    a_vilijan.predmeti.add(a_teh,a_P1)
    #a_narvika = Ucitelj(ime = "Narvika", priimek = "Bovcon", email = "narvika.bavcon@fri.uni-lj.si"))
    a_narvika = Ucitelj.objects.get(ime="Narvika",priimek="Bovcon")
    a_narvika.predmeti.add(a_obl)
    a_darja = Ucitelj(ime = "Darja", priimek = "Peljhan", email = "darja.peljhan@fri.uni-lj.si")#ep
    a_darja.save()
    a_jaka = Ucitelj(ime = "Jaka", priimek = "Lindič", email = "jaka.lindic@fri.uni-lj.si")#ep
    a_jaka.save()
    #a_mateja = Ucitelj(ime = "Mateja", priimek = "Drnovšek", email = "mateja.drnovsek@fri.uni-lj.si") #ep
    a_mateja = Ucitelj.objects.get(ime="Mateja",priimek="Drnovšek")
    a_mateja.save()
    #a = Ucitelj(ime = "Tomaž", priimek = "Hovelja", email = "tomaz.hovelja@fri.uni-lj.si")
    #a.save()
    #a = Ucitelj(ime = "Boštjan", priimek = "Slivnik", email = "bostjan.slivnik@fri.uni-lj.si")
    #a.save()
    #a = Ucitelj(ime = "Igor", priimek = "Kononenko", email = "igor.kononenko@fri.uni-lj.si")
    #a.save()

    a_izv_teh = IzvedbaPredmeta(predmet = a_teh, studijsko_leto = a_17_18, ucitelj_1 = a_vilijan)
    a_izv_teh.save()

    a_izv_teh_old = IzvedbaPredmeta(predmet = a_teh, studijsko_leto = a_16_17, ucitelj_1 = a_vilijan)

    a_izv_teh_old.save()
    #print("LO AND BEHOLD, STARA IZVEDBA IMA LETO", a_izv_teh_old.studijsko_leto)

    a = IzvedbaPredmeta(predmet = a_obl, studijsko_leto = a_17_18, ucitelj_1 = a_narvika)
    a.save()
    a = IzvedbaPredmeta(predmet = a_ep, studijsko_leto = a_17_18, ucitelj_1 = a_darja, ucitelj_2 = a_jaka, ucitelj_3 = a_mateja)
    a.save()
    
    a = Posta(id=1231, kraj="Ljubljana-Črnuče")
    a.save()
    a = Posta(id=1215, kraj="Medvode")
    a.save()



    a_stud = StudijskiProgram(id=1000475,sifra="L2",stopnja="C - (predbolonjski) univerzitetni", semestri=9, naziv= "RAČUNAL. IN INFORMATIKA UN")
    a_stud.save()
    a_stud2 = StudijskiProgram(id=1000471,sifra="L1",stopnja="L - druga stopnja: magistrski", semestri=4, naziv= "RAČUNALN. IN INFORM. MAG II.ST")
    a_stud2.save()
    a_uni = StudijskiProgram(id=1000468 ,sifra="VT",stopnja="K - prva stopnja: univerzitetni", semestri=6, naziv= "RAČUNALN. IN INFORM. UN-I.ST")
    a_uni.save()
    a_vs = StudijskiProgram(id=1000470 ,sifra="VU",stopnja="J - prva stopnja: visokošolski strokovni", semestri=6, naziv= "RAČUNALN. IN INFORM. VS-I.ST")
    a_vs.save()

    global a_vs1
    global a_vs2
    a_vs1 = VrstaStudija(id=16203,opis="Visokošolska strokovna izobrazba (prva bolonjska stopnja)", nacin_zakljucka="diplomirani...(VS)/diplomirana", raven_klasius="6/2")
    a_vs1.save()
    a_vs2 = VrstaStudija(id=16204,opis="Visokošolska univerzitetna izobrazba (prva bolonjska stopnja)", nacin_zakljucka="diplomirani...(UN)/diplomirana..(UN)", raven_klasius="6/2")
    a_vs2.save()
    a = VrstaStudija(id=17003,opis="Magistrska izobrazna (druga bolonjska stopnja)", nacin_zakljucka="magister / magistirca", raven_klasius="7")
    a.save()

    global a_vv1
    global a_vv2
    a_vv1 = VrstaVpisa(id=1, opis="Prvi vpis v letnik/dodatno leto", mozni_letniki="Vsi letniki in dodatno leto")
    a_vv1.save()
    a_vv2 = VrstaVpisa(id=2, opis="Ponavljanje letnika", mozni_letniki="V zadnjem letniku in v dodatnem letu ponavljanje ni več možno.")
    a_vv2.save()
    a = VrstaVpisa(id=3, opis="Nadaljevanje letnika", mozni_letniki="Vpis ni več dovoljen.")
    a.save()
    a = VrstaVpisa(id=4, opis="Podalšanje statusa študenta", mozni_letniki="Vsi letniki, dodatno leto")
    a.save()
    a = VrstaVpisa(id=5, opis="Vpis v semester skupnega št. programa", mozni_letniki="Vsi letniki razen prvega, dodatno leto ni dovoljeno")
    a.save()
    a = VrstaVpisa(id=6, opis="Nadaljevanje letnika", mozni_letniki="Vsi letniki, samo za skupne študijske programe")
    a.save()
    a = VrstaVpisa(id=7, opis="Nadaljevanje letnika", mozni_letniki="Vsi letniki, dodatno letno ni dovoljeno")
    a.save()
    a = VrstaVpisa(id=98, opis="Nadaljevanje letnika", mozni_letniki="Zadnji letnik. Namenjeno samo strokovmim delavcem v študentskem referatu")
    a.save()

    global a_ns1

    a_ns1 = NacinStudija(id=1, opis="redni",ang_opis="full-time")
    a_ns1.save()
    a_ns2 = NacinStudija(id=2, opis="izredni",ang_opis="part-time")
    a_ns2.save()
    #naredi 2 zetona za studenta
    global a_oblika
    a_oblika = OblikaStudija(id=1, opis="na lokaciji", ang_opis="on-site" )
    a_oblika.save()

    zeton = Zeton(student=primozt,
    studijski_program = a_uni,
    letnik=a_2Letnik,
    vrsta_vpisa=a_vv2,
    nacin_studija=a_ns1,
    vrsta_studija=a_vs2, 
    oblika_studija=a_oblika)

    zeton.save()
    zeton2 = Zeton(student=primozt,
    studijski_program=a_uni,
    letnik=a_3Letnik,
    vrsta_vpisa=a_vv1,
    nacin_studija=a_ns1,
    vrsta_studija=a_vs2, 
    oblika_studija=a_oblika)
    zeton2.save()

    zeton = Zeton(student=martin,studijski_program=StudijskiProgram.objects.filter(pk=1000468)[0],letnik=a_3Letnik,vrsta_vpisa=a_vv1,nacin_studija=a_ns1,vrsta_studija=a_vs2, oblika_studija=a_oblika)
    zeton.save()

    zeton = Zeton(student=tine,studijski_program=StudijskiProgram.objects.filter(pk=1000468)[0],letnik=a_3Letnik,vrsta_vpisa=a_vv1,nacin_studija=a_ns1,vrsta_studija=a_vs2, pravica_do_izbire = True, oblika_studija=a_oblika)
    zeton.save()

    zeton = Zeton(student=drago,studijski_program=StudijskiProgram.objects.filter(pk=1000468)[0],letnik=a_2Letnik,vrsta_vpisa=a_vv1,nacin_studija=a_ns1,vrsta_studija=a_vs2, oblika_studija=a_oblika)
    zeton.save()

    zeton = Zeton(student=ivan,studijski_program=StudijskiProgram.objects.filter(pk=1000468)[0],letnik=a_1Letnik,vrsta_vpisa=a_vv1,nacin_studija=a_ns1,vrsta_studija=a_vs2, oblika_studija=a_oblika)
    zeton.save()

    #izvedba
    a_ns2.save()
    a = IzvedbaPredmeta(predmet = a_P1, studijsko_leto = a_17_18, ucitelj_1 = a_vilijan)
    a.save()

    a_aljaz = Student(vpisna_stevilka = "63150255", emso = "5869362456789", priimek="Rupar", ime="Aljaž", naslov_stalno_bivalisce="Godešič 163, 4220 Škofja Loka", drzava=a_slo, drzava_rojstva = a_slo, posta= a_sklPosta, obcina=a_sklObcina,obcina_rojstva= a_ljObcina, telefon="031866686", email="ar1961@student.uni-lj.si")
    a_aljaz.save()

    a_stilar = Student(vpisna_stevilka = "63150253", emso = "5869362459726", priimek="Šime", ime="Štilar", naslov_stalno_bivalisce="Reteče 17, 4220 Škofja Loka", drzava=a_slo, drzava_rojstva = a_slo, posta= a_sklPosta, obcina=a_sklObcina,obcina_rojstva= a_ljObcina, telefon="031347867", email="ss1956@student.uni-lj.si")
    a_stilar.save()


    a_verlic = Student(vpisna_stevilka = "63150256", emso = "5869362456755", priimek="Verlič", ime="Aljaž", naslov_stalno_bivalisce="Voje 55, 1290 Grosuplje", drzava=a_slo,drzava_rojstva=a_slo,obcina_rojstva= a_ljObcina, posta= a_ljPosta, obcina=a_ljObcina, telefon="041786345", email="av1974@student.uni-lj.si")
    a_verlic.save()

    a_sega = Student(vpisna_stevilka = "63150245", emso = "5869362469812", priimek="Šega", ime="Nejc", naslov_stalno_bivalisce="Reteče 33, 4220 Škofja Loka", drzava=a_slo,drzava_rojstva=a_slo,obcina_rojstva= a_ljObcina, posta= a_ljPosta, obcina=a_ljObcina, telefon="041568324", email="sn1944@student.uni-lj.si")
    a_sega.save()

    a_maja = Student(vpisna_stevilka = "63150211", emso = "5869362464876", priimek="Šega", ime="Maja", naslov_stalno_bivalisce="Sveti duh, 4220 Škofja Loka", drzava=a_slo,drzava_rojstva=a_slo,obcina_rojstva= a_ljObcina, posta= a_ljPosta, obcina=a_ljObcina, telefon="041356765", email="sm1944@student.uni-lj.si")
    a_maja.save()

    a = OblikaStudija(id=1, opis="na lokaciji", ang_opis="on-site" )
    a.save()
    a = OblikaStudija(id=2, opis="na daljavo", ang_opis="distance learning" )
    a.save()
    a = OblikaStudija(id=3, opis="e-studij", ang_opis="e-learning" )
    a.save()


    a_vpisAljaz = Vpis(student=a_aljaz, studijsko_leto=a_17_18, studijski_program=a_studijskiProgram, letnik=a_3Letnik, vrsta_vpisa=a_vrstaVpisa,nacin_studija=a_nacinStudija, vrsta_studija=a_vrstaStudija)
    a_vpisAljaz.save()

    a_vpisVerlic = Vpis(student=a_verlic, studijsko_leto=a_17_18, studijski_program=a_studijskiProgram, letnik=a_3Letnik, vrsta_vpisa=a_vrstaVpisa,nacin_studija=a_nacinStudija, vrsta_studija=a_vrstaStudija)
    a_vpisVerlic.save()

    a_vpisSega = Vpis(student=a_sega, studijsko_leto=a_17_18, studijski_program=a_studijskiProgram, letnik=a_3Letnik, vrsta_vpisa=a_vrstaVpisa,nacin_studija=a_nacinStudija, vrsta_studija=a_vrstaStudija)
    a_vpisSega.save()
    
    a_vpisStilar = Vpis(student=a_stilar, studijsko_leto=a_17_18, studijski_program=a_studijskiProgram, letnik=a_3Letnik, vrsta_vpisa=a_vrstaVpisa,nacin_studija=a_nacinStudija, vrsta_studija=a_vrstaStudija)
    a_vpisStilar.save()
    
    a_vpisMaja = Vpis(student=a_maja, studijsko_leto=a_17_18, studijski_program=a_studijskiProgram, letnik=a_3Letnik, vrsta_vpisa=a_vrstaVpisa,nacin_studija=a_nacinStudija, vrsta_studija=a_vrstaStudija)
    a_vpisMaja.save()


    #a_aljaz.vpisi.add(a_vpisAljaz)
  
    a_predmetiStudentaAljaz = PredmetiStudenta()
    a_predmetiStudentaAljaz.save()
    a_predmetiStudentaAljaz.vpis = a_vpisAljaz
    a_predmetiStudentaAljaz.predmeti.add(a_teh,a_oim,a_ep,a_obl,a_P1)
    a_predmetiStudentaAljaz.save()

    a_predmetiStudentaVerlic = PredmetiStudenta()
    a_predmetiStudentaVerlic.save()
    a_predmetiStudentaVerlic.vpis = a_vpisVerlic
    a_predmetiStudentaVerlic.predmeti.add(a_teh,a_oim,a_ep,a_obl,a_aps1)
    a_predmetiStudentaVerlic.save()

    a_predmetiStudentaSega = PredmetiStudenta()
    a_predmetiStudentaSega.save()
    a_predmetiStudentaSega.vpis = a_vpisSega
    a_predmetiStudentaSega.predmeti.add(a_teh,a_oim,a_ep,a_obl,a_aps1)
    a_predmetiStudentaSega.save()
    a_predmetiStudentaStilar = PredmetiStudenta()
    a_predmetiStudentaStilar.save()
    a_predmetiStudentaStilar.vpis = a_vpisStilar
    a_predmetiStudentaStilar.predmeti.add(a_teh,a_oim,a_ep,a_obl,a_aps1)
    a_predmetiStudentaStilar.save()

    a_predmetiStudentaMaja = PredmetiStudenta()
    a_predmetiStudentaMaja.save()
    a_predmetiStudentaMaja.vpis = a_vpisMaja
    a_predmetiStudentaMaja.predmeti.add(a_teh,a_oim,a_ep,a_obl,a_aps1)
    a_predmetiStudentaMaja.save()



    new_date = datetime.datetime(2018, 2, 2, 11, 13)
    a = Rok(izvedba_predmeta = a_izv_teh, datum = new_date, prostor_izvajanja = "A1")
    a.save()

    #podatki za vpis ocen izpita
    new_date = datetime.datetime(2018, 3, 3, 14, 15)
    a_rok = Rok(izvedba_predmeta = a_izv_teh, datum = new_date, prostor_izvajanja = "A1")
    a_rok.save()

    #roki za demonstracijo
    rok_date = datetime.datetime(2018, 5, 11, 12, 10)
    rok2 = Rok(izvedba_predmeta = a_izv_teh, datum = rok_date, prostor_izvajanja = "A1")
    rok2.save()

    rok_date = datetime.datetime(2018, 5, 14, 15, 55)
    rok3 = Rok(izvedba_predmeta = a_izv_teh, datum = rok_date, prostor_izvajanja = "A2")
    rok3.save()

    # rok_date = datetime.datetime(2018, 5, 15, 6, 5)
    # rok4 = Rok(izvedba_predmeta = a_izv_teh, datum = rok_date, prostor_izvajanja = "A4")
    # rok4.save()

    # rok_date = datetime.datetime(2018, 5, 25, 13, 2)
    # rok5 = Rok(izvedba_predmeta = a_izv_teh, datum = rok_date, prostor_izvajanja = "PR08")
    # rok5.save()

    # rok_date = datetime.datetime(2018, 5, 6, 18, 7)
    # rok6 = Rok(izvedba_predmeta = a_izv_teh, datum = rok_date, prostor_izvajanja = "P16")
    # rok6.save()

    # rok_date = datetime.datetime(2018, 6, 5, 16, 35)
    # rok7 = Rok(izvedba_predmeta = a_izv_teh, datum = rok_date, prostor_izvajanja = "B")
    # rok7.save()

    # rok_date = datetime.datetime(2018, 6, 6, 12, 15)
    # rok1 = Rok(izvedba_predmeta = a_izv_teh, datum = rok_date, prostor_izvajanja = "A1")
    # rok1.save()

    # rok_date = datetime.datetime(2018, 6, 17, 12, 15)
    # rok8 = Rok(izvedba_predmeta = a_izv_teh, datum = rok_date, prostor_izvajanja = "A1")
    # rok8.save()

    # rok_date = datetime.datetime(2018, 6, 30, 12, 15)
    # rok9 = Rok(izvedba_predmeta = a_izv_teh, datum = rok_date, prostor_izvajanja = "A1")
    # rok9.save()

    # rok_date = datetime.datetime(2018, 7, 8, 12, 15)
    # rok10 = Rok(izvedba_predmeta = a_izv_teh, datum = rok_date, prostor_izvajanja = "A1")
    # rok10.save()

    # rok_date = datetime.datetime(2018, 7, 16, 12, 15)
    # rok11 = Rok(izvedba_predmeta = a_izv_teh, datum = rok_date, prostor_izvajanja = "A1")
    # rok11.save()

    # rok_date = datetime.datetime(2018, 7, 30, 12, 15)
    # rok12 = Rok(izvedba_predmeta = a_izv_teh, datum = rok_date, prostor_izvajanja = "A1")
    # rok12.save()



    new_date = datetime.datetime(2018, 2, 15, 14, 30)
    a_prijava1 = Prijava(created_at = new_date, predmeti_studenta = a_predmetiStudentaAljaz, rok = a_rok, zaporedna_stevilka_polaganja = 1)
    a_prijava1.save()

    new_date = datetime.datetime(2018, 2, 16, 14, 20)
    a_prijava2 = Prijava(created_at = new_date, predmeti_studenta = a_predmetiStudentaVerlic, rok = a_rok, zaporedna_stevilka_polaganja = 2)
    a_prijava2.save()

    new_date = datetime.datetime(2018, 2, 12, 14, 20)
    a_prijava3 = Prijava(created_at = new_date, predmeti_studenta = a_predmetiStudentaSega, rok = a_rok, zaporedna_stevilka_polaganja = 3)
    a_prijava3.save()
    new_date = datetime.datetime(2018, 2, 17, 14, 20)
    a_prijava3 = Prijava(created_at = new_date, predmeti_studenta = a_predmetiStudentaStilar, rok = a_rok, zaporedna_stevilka_polaganja = 1)
    a_prijava3.save()

    new_date = datetime.datetime(2018, 2, 12, 19, 20)
    a_prijava4 = Prijava(created_at = new_date, predmeti_studenta = a_predmetiStudentaMaja, rok = a_rok, zaporedna_stevilka_polaganja = 1)
    a_prijava4.save()



    #naredi referenta
    user, created = User.objects.get_or_create(username="referentka", email="referentka@fri.uni-lj.si")
    user.first_name = "Tatjana"
    user.last_name = "Novak"
    user.set_password("adminadmin")
    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        ref_group, status = Group.objects.get_or_create(name='referent') 
        ref_group.user_set.add(user)

    user.save()

    #naredi profesorja
    user, created = User.objects.get_or_create(username="profesor", email="profesor@fri.uni-lj.si")
    user.first_name = "Lado"
    user.last_name = "Gubara"
        
    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        prof_group, status = Group.objects.get_or_create(name='professors') 
        prof_group.user_set.add(user)

    user.save()

    user, created = User.objects.get_or_create(username="viljanmahnic", email="viljan.mahnic@fri.uni-lj.si")
    user.first_name = "Viljan"
    user.last_name = "Mahnic"
        
    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        prof_group, status = Group.objects.get_or_create(name='professors') 
        prof_group.user_set.add(user)

    user.save()

    user, created = User.objects.get_or_create(username="aljazrupar", email="ar1961@student.uni-lj.si")
    user.first_name = "Aljaž"
    user.last_name = "Rupar"
        
    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        prof_group, status = Group.objects.get_or_create(name='students') 
        prof_group.user_set.add(user)

    user.save()

    #student z absulventom za zagovor tiskanja potridl o vpisu
    #-------
    zanM = Student(vpisna_stevilka = 63150467, emso=1511996500115, ime="Žan", drzava_rojstva=a_slo,obcina_rojstva= a_ljObcina,  priimek="Mongus",naslov_stalno_bivalisce="Plečnikova cesta 12", drzava= Drzava.objects.filter(pk=4)[0], posta= Posta.objects.filter(pk=1293)[0],obcina= Obcina.objects.filter(pk=1)[0],telefon="031874563",email="zm1971@student.uni-lj.si")
    zanM.save()
    vpis_zanM = Vpis(student=zanM, studijsko_leto=a_17_18, studijski_program=a_studijskiProgram, letnik=a_3Letnik, vrsta_vpisa=a_vv1,nacin_studija=a_nacinStudija, vrsta_studija=a_vrstaStudija)
    vpis_zanM.save()


    user, created = User.objects.get_or_create(username="zanmongus", email="zm1971@student.uni-lj.si")
    user.first_name = "Žan"
    user.last_name = "Mongus"
        
    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        prof_group, status = Group.objects.get_or_create(name='students') 
        prof_group.user_set.add(user)

    user.save()


 #
    UNI = StudijskiProgram.objects.get(id=1000468)
    #LETO = StudijskoLeto.objects.get(ime="2018/2019")
    #-------
    # TESTNI PRIMER ZA KARTOTECNI LIST
    #kreiranje studenta
    print("KREIRANJE STUDENTOV")

    # 1 LETNIK
    
    student1 = naredi_studenta("Tina","Šilc","63172001","ts2001@fri.uni-lj.si")
    student2 = naredi_studenta("Lucija","Suhodolnik","63172002","ls2001@fri.uni-lj.si")
    student3 = naredi_studenta("Miha","Černe","63172003","mc20003@fri.uni-lj.si")
    student4 = naredi_studenta("Luka","Cajter","63172004","lc2004@fri.uni-lj.si")
    student5 = naredi_studenta("Samo","Sever","63172005","ss2005@fri.uni-lj.si")

    # 2 LETNIK

    #ŠPELA DOBI ŽETON ZA 2 LETNIK
    student6 = naredi_studenta("Špela","Kuhar","63162006","sk2006@fri.uni-lj.si")
    student7 = naredi_studenta("Ragnar","Lothbrok","63162007","rl2007@fri.uni-lj.si")
    student8 = naredi_studenta("Lagerta","Lothbrok","63162008","ll2008@fri.uni-lj.si")
    student9 = naredi_studenta("Sansa","Zupančič","63162009","st2009@fri.uni-lj.si")
    student10 = naredi_studenta("Klemen","Špeh","63162010","ks2010@fri.uni-lj.si")

    # 3 LETNIK
    #KLEMEN DOBI ŽETON ZA PROSTO 
    #KLEMEN IZPIŠE SVOJ KARTOTEČNI LIST
    #RAGNAR DOBI ŽETON ZA NEPROSTO IZBIRO

    student11 = naredi_studenta("Klemen","Sever","63152011","ks2011@fri.uni-lj.si")
    student12 = naredi_studenta("Branko","Pirnat","63152012","bp2012@fri.uni-lj.si")
    student13 = naredi_studenta("Tessa","Šilc","63152013","ts2013@fri.uni-lj.si")

    global student14
    global student15

    student14 = naredi_studenta("Nuša","Junhart","63152014","nj2014@fri.uni-lj.si")
    student15 = naredi_studenta("Vanesa","Novak","63152015","vs2015@fri.uni-lj.si")
    

    global student16
    global student17
    global student18
    global student19
    global student20

    student16 = naredi_studenta("Melisa","Krajcar","63152016","mk2016@fri.uni-lj.si")
    student17 = naredi_studenta("Bojan","Klemenčič","63152017","bk2017@fri.uni-lj.si")
    student18 = naredi_studenta("Viktor","Rutar","63152018","vr2018@fri.uni-lj.si")
    student19 = naredi_studenta("Dimitri","Zakeav","63152019","dz2019@fri.uni-lj.si")
    student20 = naredi_studenta("Miha","Vidmar","63152020","mv2020@fri.uni-lj.si")
    #student15 = naredi_studenta("Vanesa","Novak","63152015","vs2015@fri.uni-lj.si")

    #vpis v prvi letnik

    vpisi_studenta(student1, a_17_18, a_VS ,a_1Letnik, a_vs1)
    vpisi_studenta(student2, a_17_18, a_VS ,a_1Letnik, a_vs1)
    vpisi_studenta(student3, a_17_18, a_VS ,a_1Letnik, a_vs1)
    vpisi_studenta(student4, a_17_18, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student5, a_17_18, UNI ,a_1Letnik, a_vs2)

    vpisi_studenta(student6, a_16_17, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student7, a_16_17, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student8, a_16_17, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student9, a_16_17, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student10, a_16_17, UNI ,a_1Letnik, a_vs2)
    
    vpisi_studenta(student11, a_15_16, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student12, a_15_16, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student13, a_15_16, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student14, a_15_16, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student15, a_15_16, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student16, a_15_16, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student17, a_15_16, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student18, a_15_16, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student19, a_15_16, UNI ,a_1Letnik, a_vs2)
    vpisi_studenta(student20, a_15_16, UNI ,a_1Letnik, a_vs2)

    #vpis v drugi letnik
    #vpisi_studenta(student6, a_17_18, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student7, a_17_18, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student8, a_17_18, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student9, a_17_18, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student10, a_17_18, UNI ,a_2Letnik, a_vs2)

    vpisi_studenta(student11, a_16_17, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student12, a_16_17, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student13, a_16_17, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student14, a_16_17, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student15, a_16_17, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student16, a_16_17, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student17, a_16_17, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student18, a_16_17, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student19, a_16_17, UNI ,a_2Letnik, a_vs2)
    vpisi_studenta(student20, a_16_17, UNI ,a_2Letnik, a_vs2)
    
    # vpis v tretji letnik

    #vpisi_studenta(student11, a_17_18, UNI ,a_3Letnik, a_vs2)
    #vpisi_studenta(student12, a_17_18, UNI ,a_3Letnik, a_vs2)
    vpisi_studenta(student13, a_17_18, UNI ,a_3Letnik, a_vs2)
    vpisi_studenta(student14, a_17_18, UNI ,a_3Letnik, a_vs2)
    vpisi_studenta(student15, a_17_18, UNI ,a_3Letnik, a_vs2)
    vpisi_studenta(student16, a_17_18, UNI ,a_3Letnik, a_vs2)
    vpisi_studenta(student17, a_17_18, UNI ,a_3Letnik, a_vs2)
    vpisi_studenta(student18, a_17_18, UNI ,a_3Letnik, a_vs2)
    vpisi_studenta(student19, a_17_18, UNI ,a_3Letnik, a_vs2)
    vpisi_studenta(student20, a_17_18, UNI ,a_3Letnik, a_vs2)
    


    katarina = Student(vpisna_stevilka = "63158888", 
    emso = "1511996500207", 
    priimek="Listnik", 
    ime="Katarina", 
    naslov_stalno_bivalisce="Dunajska cesta 110", 
    drzava=a_slo, 
    drzava_rojstva = a_slo, 
    posta= a_sklPosta, 
    obcina=a_sklObcina,
    obcina_rojstva= a_ljObcina, 
    telefon="031866686", 
    email="kl9999@student.uni-lj.si")
    katarina.save()

    user, created = User.objects.get_or_create(username="katarinalistnik", email="kl9999@student.uni-lj.si")
    user.first_name = "Katarina"
    user.last_name = "Listnik"
        
    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
        prof_group, status = Group.objects.get_or_create(name='students') 
        prof_group.user_set.add(user)

    user.save()

    

    # kreiranje vpisa
    katarina_vpis = Vpis(student=katarina, 
    studijsko_leto= a_15_16, 
    studijski_program=UNI, 
    letnik=a_1Letnik, 
    vrsta_vpisa=a_vv1,
    nacin_studija=a_ns1,
    oblika_studija = a_oblika_on_site,
    potrjen = True, 
    dokoncan_vpis = True,
    vrsta_studija=a_vs2)
    katarina_vpis.save()

    katarina_vpis1_2 = Vpis(student=katarina, 
    studijsko_leto= a_16_17, 
    studijski_program=UNI, 
    letnik=a_1Letnik, 
    vrsta_vpisa=a_vv2,
    nacin_studija=a_ns1,
    oblika_studija = a_oblika_on_site,
    potrjen = True, 
    dokoncan_vpis = True,
    vrsta_studija=a_vs2)
    katarina_vpis1_2.save()

    katarina_vpis2 = Vpis(student=katarina, 
    studijsko_leto=a_17_18, 
    studijski_program=UNI,
    oblika_studija = a_oblika_on_site,
    letnik=a_2Letnik, 
    vrsta_vpisa=a_vv1,
    nacin_studija=a_ns1,
    potrjen = True,
    dokoncan_vpis = True,
    vrsta_studija=a_vs2)
    katarina_vpis2.save()

    # dodajanje predmetov
    #katartina predmeti so spelini predmeti, i know im lazy
    vpis1 = Vpis.objects.get(student = student11, studijsko_leto__ime = "2015/2016")
    katarina_predmeti = PredmetiStudenta.objects.get(vpis = vpis1 )


    # dodajanje predmetov
    vpis2 = Vpis.objects.get(student = student11, studijsko_leto__ime = "2016/2017")
    katarina_predmeti2 = PredmetiStudenta.objects.get(vpis = vpis2 )
    
    # prvi 
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[0], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016, 3, 3, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 3, 3, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 6)

    prijava_1.save()
    # drugi
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[1], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016, 4, 4, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 4, 4, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 5)

    prijava_1.save()
    # drugi drugic 
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[1], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016, 5, 5, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 5, 5, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 2, 
        ocena_izpita = 5)

    prijava_1.save()

    # drugi tretjic neuspesno 
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[1], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016, 7, 7, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 7, 7, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 3, 
        ocena_izpita = 6)

    prijava_1.save()


    #tretji uspesno
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[2], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016, 8, 8, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 8, 8, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 8)

    prijava_1.save()

    #cetrti uspesno
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[3], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016, 6, 10, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 6, 10, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 7)

    prijava_1.save()
    #peti uspesno
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[4], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016, 5, 4, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 5, 4, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 6)

    prijava_1.save()

    #sesti prvic
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[5], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016, 4, 14, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 4, 14, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 5)

    prijava_1.save()

    #sesti drugic neuspesno
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[5], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016,  5, 20, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 5, 20, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 2, 
        ocena_izpita = 8)

    prijava_1.save()

    #sedmic prvic neuspesno 

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[6], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016,  4, 20, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 4, 20, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 7)

    prijava_1.save()

    #osmic prvic  

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[7], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016,  5, 10, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 5, 10, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 5)

    prijava_1.save()
    #osmic drugic  uspesno

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[7], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016,  6, 1, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 6, 1, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 2, 
        ocena_izpita = 10)

    prijava_1.save()


    #devetic prvi

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[8], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016,  6, 1, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 6, 1, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 5)

    prijava_1.save()
    #devetic drugic 

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[8], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016,  6, 13, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 6, 13, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 2, 
        ocena_izpita = 5)

    prijava_1.save()
    #devetic tretjic neuspesno

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[8], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016,  6, 19, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 6, 19, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 3, 
        ocena_izpita = 10)

    prijava_1.save()

    #deset tretjic neuspesno

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[9], studijsko_leto = a_15_16)[0]
    new_date = datetime.datetime(2016,  6, 20, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2016, 6, 20, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti, 
        rok = rok_1, zaporedna_stevilka_polaganja = 3, 
        ocena_izpita = 10)

    prijava_1.save()

    #drugi letnik 
    #drjgi letnik
    #drugi letnik 
    #drjgi letnik
    #drugi letnik 
    #drjgi letnik
    #drugi letnik 
    #drjgi letnik

        # prvi 
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[0], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017, 3, 3, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 3, 3, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 6)

    prijava_1.save()
    # drugi
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[1], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017, 4, 4, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 4, 4, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 5)

    prijava_1.save()
    # drugi drugic 
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[1], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017, 5, 5, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 5, 5, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 2, 
        ocena_izpita = 9)

    prijava_1.save()



    #tretji uspesno
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[2], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017, 8, 8, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 8, 8, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 8)

    prijava_1.save()

    #cetrti uspesno
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[3], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017, 6, 10, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 6, 10, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 7)

    prijava_1.save()
    #peti uspesno
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[4], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017, 5, 4, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 5, 4, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 6)

    prijava_1.save()

    #sesti prvic
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[5], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017, 5, 14, 14, 15)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 5, 14, 14, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 5)

    prijava_1.save()

    #sesti drugic neuspesno
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[5], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017,  5, 20, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 5, 20, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 2, 
        ocena_izpita = 6)

    prijava_1.save()

    #sedmic prvic neuspesno 

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[6], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017,  4, 20, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 4, 20, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 9)

    prijava_1.save()

    #osmic prvic  

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[7], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017,  5, 10, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 5, 10, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 5)

    prijava_1.save()
    #osmic drugic  uspesno

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[7], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017,  6, 1, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 6, 1, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 2, 
        ocena_izpita = 8)

    prijava_1.save()


    #devetic prvi

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[8], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017,  6, 1, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 6, 1, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 1, 
        ocena_izpita = 8)

    prijava_1.save()
    #devetic drugic 


    #devetic tretjic neuspesno


    #deset tretjic neuspesno

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_DRUGI[9], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017,  6, 20, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 6, 20, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 3, 
        ocena_izpita = 9)

    prijava_1.save()
    '''
    # 1 LETNIK PONAVLJANJE
    # prvi predmet 
    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[1], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017,  3, 20, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 3, 20, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti1_2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 4, 
        ocena_izpita = 10)

    prijava_1.save()

    #drugi predmet prvic

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[5], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017,  4, 20, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 4, 20, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti1_2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 4, 
        ocena_izpita = 5)

    prijava_1.save()

    #drugi predmet drugic uspesno

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[5], studijsko_leto = a_16_17)[0]
    new_date = datetime.datetime(2017,  5, 20, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 5, 20, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti1_2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 5, 
        ocena_izpita = 8)

    prijava_1.save()

    #tretji predmet drugic uspesno

    #izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[6], studijsko_leto = a_16_17)[0]
    pr = Predmetnik.objects.filter(studijski_program = UNI).filter(studijsko_leto= a_16_17).filter(predmet= UNI_PREDMETI_PRVI[6])[0]
    narediIzvedboSTremi(pr, narediProfesorja("Nov","Profesor1"),narediProfesorja("Nov","Profesor2"),narediProfesorja("Nov","Profesor2"))
    izvedba_1 = IzvedbaPredmeta.objects.filter(ucitelj_1__priimek="Profesor1")[0]
    new_date = datetime.datetime(2017,  5, 20, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 5, 20, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti1_2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 5, 
        ocena_izpita = 8)

    prijava_1.save()

    #cetrti predmet prvic uspesno

    izvedba_1 = IzvedbaPredmeta.objects.filter(predmet=UNI_PREDMETI_PRVI[8], studijsko_leto = a_16_17)[0]

    
    new_date = datetime.datetime(2017,  7, 20, 15,30)
    rok_1 = Rok(izvedba_predmeta = izvedba_1, datum = new_date, prostor_izvajanja = "P1")
    rok_1.save()

    new_date = datetime.datetime(2017, 7, 20, 20, 15)
    prijava_1 = Prijava(created_at = new_date, 
        predmeti_studenta = katarina_predmeti1_2, 
        rok = rok_1, zaporedna_stevilka_polaganja = 5, 
        ocena_izpita = 10)

    prijava_1.save()
    '''






    print('\x1b[6;30;42m' + 'Uspešno dodana baza! ' + '\x1b[0m')
    return HttpResponse("Narejena baza!")

    

def vsi_predmeti():

    UNI = StudijskiProgram.objects.get(id=1000468)
    LETO = StudijskoLeto.objects.get(ime="2018/2019")
    LETNIK = Letnik.objects.get(ime="1.")

    #1 letnik
    a = Predmet(ime = "Programiranje 1", id="63277")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    #narediIzvedbo(pr, Ucitelj.objects.get(email="viljan.mahnic@fri.uni-lj.si"))
    narediIzvedboSTremi(pr, Ucitelj.objects.get(email="viljan.mahnic@fri.uni-lj.si"),narediProfesorja("Luka","Fuerst"), narediProfesorja("Marko","Poženel") )
    a = Predmet(ime = "Programiranje 2", id="63278")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Boštjan","Slivnik"))
    
    a = Predmet(ime = "Diskretne strukture", id="63203")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Gašper","Fijavž"))

    a = Predmet(ime = "Fizika", id="63205")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Jan","Ravnik"))

    a = Predmet(ime = "Osnove digitalnih vezij", id="63204")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Nikolaj","Zimic"))

    a = Predmet(ime = "Osnove matematične analize", id="63202")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Neža","Mramor"))

    a = Predmet(ime = "Linearna algebra", id="63207")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Bojan","Orel"))

    a = Predmet(ime = "Osnove informacijskih sistemov", id="63215")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Dejan","Lavbič"))

    a = Predmet(ime = "Računalniške komunikacije", id="63209")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Zoran","Bosnić"))

    a = Predmet(ime = "Arhitektura računalniških sistemov", id="63212")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Branko","Šter"))

    LETNIK = Letnik.objects.get(ime="2.")



    #2 letnik obvezni
    a = Predmet(ime = "Verjetnost in statistika", id="63213")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Aleksander","Jurišić"))

    a = Predmet(ime = "Algoritmi in podatkovne strukture 1", id="63279")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Igor","Kononenko"))

    a = Predmet(ime = "Osnove podatkovnih baz", id="63208")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Marko","Bajec"))

    a = Predmet(ime = "Organizacija računalniških sistemov", id="63218")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()

    robic = narediProfesorja("Borut","Robič")
    narediIzvedbo(pr, narediProfesorja("Patricio","Bulić"))

    a = Predmet(ime = "Izračunljivost in računska zahtevnost", id="63283")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr,robic )

    a = Predmet(ime = "Teorija informacij in sistemov", id="63216")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Uros","Lotric"))

    a = Predmet(ime = "Algoritmi in podatkovne strukture 2", id="63280")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, robic)

    a = Predmet(ime = "Operacijski sistemi", id="63217")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, robic)

    #2 letnik strokovni
    a = Predmet(ime = "Principi programskih jezikov", id="63220", )
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, strokoven=True)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Andrej","Bauer"))

    a = Predmet(ime = "Računalniške tehnologije", id="63221")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, strokoven=True)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Rok","Žitko"))

    a = Predmet(ime = "Matematično modeliranje", id="63219")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, strokoven=True)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Neža","Mramor"))

    LETNIK = Letnik.objects.get(ime="3.")

    
    

    #3 letnik obvezni
    a = Predmet(ime = "Osnove umetne inteligence", id="63214")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Zoran","Bosnić"))

    a = Predmet(ime = "Ekonomika in podjetništvo", id="63248")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Mateja","Drnovšek"))

    a = Predmet(ime = "Diplomski seminar", id="63281")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Franc","Solina"))


    #moduli
    modul1 = Modul(ime="Informacijski sistemi", studijsko_leto=LETO, studijski_program=UNI)
    modul1.save()
    modul2 = Modul(ime="Obvladovanje informatike", studijsko_leto=LETO, studijski_program=UNI)
    modul2.save()
    modul3 = Modul(ime="Računalniška omrežja", studijsko_leto=LETO, studijski_program=UNI)
    modul3.save()
    modul4 = Modul(ime="Umetna inteligenca", studijsko_leto=LETO, studijski_program=UNI)
    modul4.save()
    modul5 = Modul(ime="Razvoj programske opreme", studijsko_leto=LETO, studijski_program=UNI)
    modul5.save()
    modul6 = Modul(ime="Medijske tehnologije", studijsko_leto=LETO, studijski_program=UNI)
    modul6.save()
    modul7 = Modul(ime="Algoritmi in sistemski programi", studijsko_leto=LETO, studijski_program=UNI)
    modul7.save()

    
    

    #informacijski sistemi
    a = Predmet(ime = "Elektronsko poslovanje", id="63249")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul1)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Denis","Trček"))

    a = Predmet(ime = "Poslovna inteligenca", id="63251")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul1)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Darko","Smedlnik"))

    a = Predmet(ime = "Organizacija in management", id="63250")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul1)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Tomaž","Hovelja"))

    #obladovanje informatike
    a = Predmet(ime = "Razvoj informacijskih sistemov", id="63252")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul2)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Marko","Bajec"))

    a = Predmet(ime = "Tehnologija upravljanja podatkov", id="63226")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul2)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Matjaž","Kukar"))

    a = Predmet(ime = "Planiranje in upravljanje informatike", id="63253")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul2)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Rok","Rupnik"))


    #racunalniska omrezja
    a = Predmet(ime = "Modeliranje računalniških omrežij", id="63257")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul3)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Miha","Mraz"))

    a = Predmet(ime = "Komunikacijski protokoli", id="63258")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul3)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Mojca","Ceglarič"))

    a = Predmet(ime = "Brezžična in mobilna omrežja", id="63259")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul3)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Nikolaj","Zimic"))
    
    

    #umetna inteligenca
    a = Predmet(ime = "Inteligentni sistemi", id="63266")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul4)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Marko","Robnik"))

    a = Predmet(ime = "Umetno zaznavanje", id="63267")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul4)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Matej","Kristan"))

    a = Predmet(ime = "Razvoj inteligentnih sistemov", id="63262")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul4)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Danijel","Skočaj"))


    #razvoj programske opreme
    a = Predmet(ime = "Postopki razvoja programske opreme", id="63254")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul5)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Matjaž","Jurič"))

    a = Predmet(ime = "Spletno programiranje", id="63255")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul5)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Dejan","Lavbič"))

    a = Predmet(ime = "Tehnologija programske opreme", id="63256")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul5)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Viljan","Mahnič"))


    #medijske tehnologije
    a = Predmet(ime = "Računalniška grafika in tehnologija iger", id="63269")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul6)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Matija","Marolt"))

    a = Predmet(ime = "Multimedijski sistemi", id="63270")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul6)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Luka","Zajc"))

    a = Predmet(ime = "Osnove oblikovanja", id="63271")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul6)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Narvika","Bovcon"))

    #algoritmi in sistemski programi
    a = Predmet(ime = "Računska zahtevnost in hevristično programiranje", id="63263")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul7)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Marko","Robnik Šikonja"))

    a = Predmet(ime = "Sistemska programska oprema", id="63264")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul7)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Tomaž","Dobravec"))


    a = Predmet(ime = "Prevajalniki", id="63265")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul7)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Boštjan","Slivnik"))


    

    LETNIK = Letnik.objects.get(ime="2.")

    #splosno izbirni predmeti
    a = Predmet(ime = "Tehnične veščine", id="63284", kreditne_tocke=3)
    a.save()
    LETNIK = Letnik.objects.get(ime="2.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    #narediIzvedbo(pr, narediProfesorja("Luka","Zajc"))

    LETNIK = Letnik.objects.get(ime="3.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Luka","Zajc"))


    a = Predmet(ime = "Angleški jezik nivo A", id="63222", kreditne_tocke=3)
    a.save()
    LETNIK = Letnik.objects.get(ime="2.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    #narediIzvedbo(pr, narediProfesorja("Barbara","Slivnik"))

    LETNIK = Letnik.objects.get(ime="3.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Barbara","Slivnik"))


    a = Predmet(ime = "Angleški jezik nivo B", id="63746", kreditne_tocke=3)
    a.save()
    LETNIK = Letnik.objects.get(ime="2.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    #narediIzvedbo(pr, narediProfesorja("Barbara","Slivnik"))

    LETNIK = Letnik.objects.get(ime="3.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Barbara","Slivnik"))


    a = Predmet(ime = "Angleški jezik nivo C", id="63747", kreditne_tocke=3)
    a.save()
    LETNIK = Letnik.objects.get(ime="2.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    #narediIzvedbo(pr, narediProfesorja("Barbara","Slivnik"))

    LETNIK = Letnik.objects.get(ime="3.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Barbara","Slivnik"))

    a = Predmet(ime = "Računalništvo v praksi I", id="63752", kreditne_tocke=3)
    a.save()
    LETNIK = Letnik.objects.get(ime="2.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    #narediIzvedbo(pr, narediProfesorja("Klemen","Sekir"))
    LETNIK = Letnik.objects.get(ime="3.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Klemen","Sekir"))

    a = Predmet(ime = "Računalništvo v praksi II", id="63242", kreditne_tocke=3)
    a.save()
    LETNIK = Letnik.objects.get(ime="2.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    #narediIzvedbo(pr, narediProfesorja("Klemen","Sekir"))
    LETNIK = Letnik.objects.get(ime="3.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Klemen","Sekir"))

    VSS = StudijskiProgram.objects.get(id=1000470)
    #LETO = StudijskoLeto.objects.get(ime="2018/2019")
    LETNIK = Letnik.objects.get(ime="1.")

    # prvi letnik vss

    a = Predmet(ime = "Osnove verjetnosti in statistike", id="63710")
    a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Aleksander","Jurišić"))

    a = Predmet(ime = "Operacijski sistemi", id="63709")
    a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Peter","Peer"))
    
    a = Predmet(ime = "Računalniške komunikacije", id="63708")
    a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Mojca","Ciglarič"))

    a = Predmet(ime = "Programiranje 1", id="63702")
    a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Janez","Demšar"))

    a = Predmet(ime = "Podatkovne baze", id="63707")
    a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Matjaž","Kukar"))

    a = Predmet(ime = "Programiranje 2", id="63706")
    a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Tomaž","Dobravec"))

    a = Predmet(ime = "Diskretne strukture", id="63705")
    a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Sandi","Klavžar"))

    a = Predmet(ime = "Matematika", id="63704")
    a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Roman","Drnovšek"))

    a = Predmet(ime = "Uvod v računalništvo", id="63701")
    a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Danijel","Skočaj"))

    a = Predmet(ime = "Računalniška arhitektura", id="63703")
    a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Igor","Škraba"))
##########################################
#
# vsi_predmeti2 se razlikujejo z vsi_predmeti le v studeijskem letu
# ne kreira ponovno predmetov!
#
###########################################
def vsi_predmeti2(LETO):

    UNI = StudijskiProgram.objects.get(id=1000468)
    #LETO = StudijskoLeto.objects.get(ime="2017/2018")
    LETNIK = Letnik.objects.get(ime="1.")

    #1 letnik
    a = Predmet.objects.get(ime = "Programiranje 1", id="63277")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedboSTremi(pr, Ucitelj.objects.get(email="viljan.mahnic@fri.uni-lj.si"),narediProfesorja("Luka","Fuerst"), narediProfesorja("Marko","Poženel") )
    a = Predmet.objects.get(ime = "Programiranje 2", id="63278")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Boštjan","Slivnik"))

    a = Predmet.objects.get(ime = "Diskretne strukture", id="63203")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Gašper","Fijavž"))

    a = Predmet.objects.get(ime = "Fizika", id="63205")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Jan","Ravnik"))

    a = Predmet.objects.get(ime = "Osnove digitalnih vezij", id="63204")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Nikolaj","Zimic"))

    a = Predmet.objects.get(ime = "Osnove matematične analize", id="63202")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Neža","Mramor"))

    a = Predmet.objects.get(ime = "Linearna algebra", id="63207")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Bojan","Orel"))

    a = Predmet.objects.get(ime = "Osnove informacijskih sistemov", id="63215")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Dejan","Lavbič"))

    a = Predmet.objects.get(ime = "Računalniške komunikacije", id="63209")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Zoran","Bosnić"))

    a = Predmet.objects.get(ime = "Arhitektura računalniških sistemov", id="63212")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Branko","Šter"))

    LETNIK = Letnik.objects.get(ime="2.")



    #2 letnik obvezni
    a = Predmet.objects.get(ime = "Verjetnost in statistika", id="63213")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Aleksander","Jurišić"))

    a = Predmet.objects.get(ime = "Algoritmi in podatkovne strukture 1", id="63279")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Igor","Kononenko"))

    a = Predmet.objects.get(ime = "Osnove podatkovnih baz", id="63208")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Marko","Bajec"))

    a = Predmet.objects.get(ime = "Organizacija računalniških sistemov", id="63218")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()

    robic = narediProfesorja("Borut","Robič")
    narediIzvedbo(pr, narediProfesorja("Patricio","Bulić"))

    a = Predmet.objects.get(ime = "Izračunljivost in računska zahtevnost", id="63283")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr,robic )

    a = Predmet.objects.get(ime = "Teorija informacij in sistemov", id="63216")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Uros","Lotric"))

    a = Predmet.objects.get(ime = "Algoritmi in podatkovne strukture 2", id="63280")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, robic)

    a = Predmet.objects.get(ime = "Operacijski sistemi", id="63217")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, robic)

    #2 letnik strokovni
    a = Predmet.objects.get(ime = "Principi programskih jezikov", id="63220", )
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, strokoven=True)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Andrej","Bauer"))

    a = Predmet.objects.get(ime = "Računalniške tehnologije", id="63221")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, strokoven=True)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Rok","Žitko"))

    a = Predmet.objects.get(ime = "Matematično modeliranje", id="63219")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, strokoven=True)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Neža","Mramor"))
    
    LETNIK = Letnik.objects.get(ime="3.")

    #3 letnik obvezni
    a = Predmet.objects.get(ime = "Osnove umetne inteligence", id="63214")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Zoran","Bosnić"))

    a = Predmet.objects.get(ime = "Ekonomika in podjetništvo", id="63248")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Mateja","Drnovšek"))

    a = Predmet.objects.get(ime = "Diplomski seminar", id="63281")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Franc","Solina"))


    #moduli
    modul1 = Modul(ime="Informacijski sistemi", studijsko_leto=LETO, studijski_program = UNI)
    modul1.save()
    modul2 = Modul(ime="Obvladovanje informatike", studijsko_leto=LETO, studijski_program = UNI)
    modul2.save()
    modul3 = Modul(ime="Računalniška omrežja", studijsko_leto=LETO, studijski_program = UNI)
    modul3.save()
    modul4 = Modul(ime="Umetna inteligenca", studijsko_leto=LETO, studijski_program = UNI)
    modul4.save()
    modul5 = Modul(ime="Razvoj programske opreme", studijsko_leto=LETO, studijski_program = UNI)
    modul5.save()
    modul6 = Modul(ime="Medijske tehnologije", studijsko_leto=LETO, studijski_program = UNI)
    modul6.save()
    modul7 = Modul(ime="Algoritmi in sistemski programi", studijsko_leto=LETO, studijski_program = UNI)
    modul7.save()

    #informacijski sistemi
    a = Predmet.objects.get(ime = "Elektronsko poslovanje", id="63249")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul1)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Denis","Trček"))

    a = Predmet.objects.get(ime = "Poslovna inteligenca", id="63251")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul1)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Darko","Smedlnik"))

    a = Predmet.objects.get(ime = "Organizacija in management", id="63250")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul1)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Tomaž","Hovelja"))

    #obladovanje informatike
    a = Predmet.objects.get(ime = "Razvoj informacijskih sistemov", id="63252")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul2)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Marko","Bajec"))

    a = Predmet.objects.get(ime = "Tehnologija upravljanja podatkov", id="63226")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul2)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Matjaž","Kukar"))

    a = Predmet.objects.get(ime = "Planiranje in upravljanje informatike", id="63253")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul2)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Rok","Rupnik"))


    #racunalniska omrezja
    a = Predmet.objects.get(ime = "Modeliranje računalniških omrežij", id="63257")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul3)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Miha","Mraz"))

    a = Predmet.objects.get(ime = "Komunikacijski protokoli", id="63258")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul3)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Mojca","Ceglarič"))

    a = Predmet.objects.get(ime = "Brezžična in mobilna omrežja", id="63259")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul3)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Nikolaj","Zimic"))


    #umetna inteligenca
    a = Predmet.objects.get(ime = "Inteligentni sistemi", id="63266")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul4)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Marko","Robnik"))

    a = Predmet.objects.get(ime = "Umetno zaznavanje", id="63267")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul4)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Matej","Kristan"))

    a = Predmet.objects.get(ime = "Razvoj inteligentnih sistemov", id="63262")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul4)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Danijel","Skočaj"))


    #razvoj programske opreme
    a = Predmet.objects.get(ime = "Postopki razvoja programske opreme", id="63254")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul5)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Matjaž","Jurič"))

    a = Predmet.objects.get(ime = "Spletno programiranje", id="63255")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul5)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Dejan","Lavbič"))

    a = Predmet.objects.get(ime = "Tehnologija programske opreme", id="63256")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul5)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Viljan","Mahnič"))


    #medijske tehnologije
    a = Predmet.objects.get(ime = "Računalniška grafika in tehnologija iger", id="63269")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul6)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Matija","Marolt"))

    a = Predmet.objects.get(ime = "Multimedijski sistemi", id="63270")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul6)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Luka","Zajc"))

    a = Predmet.objects.get(ime = "Osnove oblikovanja", id="63271")
    #a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul6)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Narvika","Bovcon"))

    #algoritmi in sistemski programi
    a = Predmet.objects.get(ime = "Računska zahtevnost in hevristično programiranje", id="63263")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul7)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Marko","Robnik Šikonja"))

    a = Predmet.objects.get(ime = "Sistemska programska oprema", id="63264")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul7)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Tomaž","Dobravec"))

    a = Predmet.objects.get(ime = "Prevajalniki", id="63265")
    a.save()
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False, modul=modul7)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Boštjan","Slivnik"))


    LETNIK = Letnik.objects.get(ime="2.")

    #splosno izbirni predmeti
    a = Predmet.objects.get(ime = "Tehnične veščine", id="63284", kreditne_tocke=3)
    #a.save()
    LETNIK = Letnik.objects.get(ime="2.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    #narediIzvedbo(pr, narediProfesorja("Luka","Zajc"))

    LETNIK = Letnik.objects.get(ime="3.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Luka","Zajc"))


    a = Predmet.objects.get(ime = "Angleški jezik nivo A", id="63222", kreditne_tocke=3)
    #a.save()
    LETNIK = Letnik.objects.get(ime="2.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    #narediIzvedbo(pr, narediProfesorja("Barbara","Slivnik"))

    LETNIK = Letnik.objects.get(ime="3.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Barbara","Slivnik"))


    a = Predmet.objects.get(ime = "Angleški jezik nivo B", id="63746", kreditne_tocke=3)
    #a.save()
    LETNIK = Letnik.objects.get(ime="2.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    #narediIzvedbo(pr, narediProfesorja("Barbara","Slivnik"))

    LETNIK = Letnik.objects.get(ime="3.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Barbara","Slivnik"))


    a = Predmet.objects.get(ime = "Angleški jezik nivo C", id="63747", kreditne_tocke=3)
    #a.save()
    LETNIK = Letnik.objects.get(ime="2.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    #narediIzvedbo(pr, narediProfesorja("Barbara","Slivnik"))

    LETNIK = Letnik.objects.get(ime="3.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Barbara","Slivnik"))

    a = Predmet.objects.get(ime = "Računalništvo v praksi I", id="63752", kreditne_tocke=3)
    #a.save()
    LETNIK = Letnik.objects.get(ime="2.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    #narediIzvedbo(pr, narediProfesorja("Klemen","Sekir"))
    LETNIK = Letnik.objects.get(ime="3.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Klemen","Sekir"))

    a = Predmet.objects.get(ime = "Računalništvo v praksi II", id="63242", kreditne_tocke=3)
    #a.save()
    LETNIK = Letnik.objects.get(ime="2.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    #narediIzvedbo(pr, narediProfesorja("Klemen","Sekir"))
    LETNIK = Letnik.objects.get(ime="3.")
    pr = Predmetnik(studijski_program = UNI, studijsko_leto=LETO, letnik = LETNIK, predmet = a, obvezen=False)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Klemen","Sekir"))

    VSS = StudijskiProgram.objects.get(id=1000470)
    #LETO = StudijskoLeto.objects.get(ime="2017/2018")
    LETNIK = Letnik.objects.get(ime="1.")

    # prvi letnik vss



    a = Predmet.objects.get(ime = "Osnove verjetnosti in statistike", id="63710")
    #a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Aleksander","Jurišić"))

    a = Predmet.objects.get(ime = "Operacijski sistemi", id="63709")
    #a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Peter","Peer"))
    
    a = Predmet.objects.get(ime = "Računalniške komunikacije", id="63708")
    #a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Mojca","Ciglarič"))

    a = Predmet.objects.get(ime = "Programiranje 1", id="63702")
    #a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Janez","Demšar"))

    a = Predmet.objects.get(ime = "Podatkovne baze", id="63707")
    #a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Matjaž","Kukar"))

    a = Predmet.objects.get(ime = "Programiranje 2", id="63706")
    #a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Tomaž","Dobravec"))

    a = Predmet.objects.get(ime = "Diskretne strukture", id="63705")
    #a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Sandi","Klavžar"))

    a = Predmet.objects.get(ime = "Matematika", id="63704")
    #a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Roman","Drnovšek"))

    a = Predmet.objects.get(ime = "Uvod v računalništvo", id="63701")
    #a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Danijel","Skočaj"))

    a = Predmet.objects.get(ime = "Računalniška arhitektura", id="63703")
    #a.save()
    pr = Predmetnik(studijski_program = VSS, studijsko_leto=LETO, letnik = LETNIK, predmet = a)
    pr.save()
    narediIzvedbo(pr, narediProfesorja("Igor","Škraba"))
    '''
    LETO = StudijskoLeto.objects.get(ime="2016/2017")
    modul1 = Modul(ime="Informacijski sistemi", studijsko_leto=LETO, studijski_program=UNI)
    modul1.save()
    modul2 = Modul(ime="Obvladovanje informatike", studijsko_leto=LETO, studijski_program=UNI)
    modul2.save()
    modul3 = Modul(ime="Računalniška omrežja", studijsko_leto=LETO, studijski_program=UNI)
    modul3.save()
    modul4 = Modul(ime="Umetna inteligenca", studijsko_leto=LETO, studijski_program=UNI)
    modul4.save()
    modul5 = Modul(ime="Razvoj programske opreme", studijsko_leto=LETO, studijski_program=UNI)
    modul5.save()
    modul6 = Modul(ime="Medijske tehnologije", studijsko_leto=LETO, studijski_program=UNI)
    modul6.save()
    modul7 = Modul(ime="Algoritmi in sistemski programi", studijsko_leto=LETO, studijski_program=UNI)
    modul7.save()
    '''
